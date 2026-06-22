"""
Excel hisobot eksporti — openpyxl asosida.

Funksiyalar:
- export_sales(month)   → sotuvlar (chek-chek)
- export_clients()      → barcha mijozlar + qarzlari
- export_debtors()      → faqat qarzdorlar
- export_products()     → mahsulotlar ombor holati
"""

import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.channel_db import db, DB_PATH, now_local


_OUTPUT_DIR = os.path.join(os.path.dirname(DB_PATH), "exports")
os.makedirs(_OUTPUT_DIR, exist_ok=True)


# ─── Stillar ─────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")


def _autosize(ws):
    """Ustun kengligini avtomatik moslaydi (merged kataklarni e'tiborsiz qoldiradi)."""
    widths: dict = {}
    for row in ws.iter_rows():
        for cell in row:
            # MergedCell uchun column_letter yo'q
            col_letter = getattr(cell, "column_letter", None)
            if not col_letter or cell.value is None:
                continue
            try:
                ln = len(str(cell.value))
                if ln > widths.get(col_letter, 0):
                    widths[col_letter] = ln
            except Exception:
                pass
    for col, ln in widths.items():
        ws.column_dimensions[col].width = min(ln + 4, 50)


def _write_header(ws, headers: list, row: int = 1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER


def _money(v) -> str:
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "0"


# ─── Sotuvlar ────────────────────────────────────────────────────────────────

async def export_sales(month: str = None) -> str:
    """
    Sotuvlar hisoboti. month bo'lsa shu oydagi, aks holda barcha sotuvlar.
    Qaytaradi: yaratilgan .xlsx fayl yo'lini.
    """
    if month:
        sales = await db.get_sales_by_month(month)
        title = f"Sotuvlar — {month}"
    else:
        sales = await db.get_all_sales()
        title = "Barcha sotuvlar"

    wb = Workbook()

    # --- 1-varaq: Umumiy ro'yxat ---
    ws = wb.active
    ws.title = "Sotuvlar"
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:H1")

    headers = ["#", "Sana", "Kassir", "Jami so'm", "Jami USD", "Kurs",
               "Naqd", "Karta", "Nasiya", "Mijoz"]
    _write_header(ws, headers, row=3)

    total_sum = 0.0
    total_usd = 0.0
    cash_sum = card_sum = nasiya_sum = 0.0

    for i, s in enumerate(sales, 1):
        r = i + 3
        ws.cell(row=r, column=1, value=s["id"]).alignment = _CENTER
        ws.cell(row=r, column=2, value=s["created_at"][:16])
        ws.cell(row=r, column=3, value=s.get("cashier_name", ""))
        c4 = ws.cell(row=r, column=4, value=float(s["total"]))
        c4.number_format = '#,##0'
        c4.alignment = _RIGHT
        c5 = ws.cell(row=r, column=5, value=float(s.get("total_usd", 0) or 0))
        c5.number_format = '$#,##0.00'
        c5.alignment = _RIGHT
        c6 = ws.cell(row=r, column=6, value=float(s.get("usd_rate", 0) or 0))
        c6.number_format = '#,##0.00'
        c7 = ws.cell(row=r, column=7, value=float(s.get("paid_cash", 0)))
        c7.number_format = '#,##0'
        c8 = ws.cell(row=r, column=8, value=float(s.get("paid_card", 0)))
        c8.number_format = '#,##0'
        c9 = ws.cell(row=r, column=9,
                     value=float(s["total"]) if s.get("is_nasiya") else 0)
        c9.number_format = '#,##0'
        ws.cell(row=r, column=10, value=s.get("client_name", ""))

        for col in range(1, 11):
            ws.cell(row=r, column=col).border = _BORDER

        total_sum += s["total"]
        total_usd += s.get("total_usd", 0) or 0
        cash_sum += s.get("paid_cash", 0)
        card_sum += s.get("paid_card", 0)
        if s.get("is_nasiya"):
            nasiya_sum += s["total"]

    # Yakuniy qator
    last = len(sales) + 4
    ws.cell(row=last, column=3, value="JAMI:").font = Font(bold=True)
    for col, val, fmt in [
        (4, total_sum, '#,##0'),
        (5, total_usd, '$#,##0.00'),
        (7, cash_sum, '#,##0'),
        (8, card_sum, '#,##0'),
        (9, nasiya_sum, '#,##0'),
    ]:
        c = ws.cell(row=last, column=col, value=val)
        c.font = Font(bold=True)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    _autosize(ws)

    # --- 2-varaq: Tafsilotlar (har bir mahsulot) ---
    ws2 = wb.create_sheet("Tafsilot")
    ws2["A1"] = f"{title} — Mahsulot tafsiloti"
    ws2["A1"].font = _TITLE_FONT
    ws2.merge_cells("A1:G1")

    _write_header(
        ws2,
        ["Sotuv #", "Sana", "Mahsulot", "Miqdori", "Birligi",
         "Narx (so'm)", "Narx (USD)", "Jami so'm", "Jami USD"],
        row=3
    )
    rr = 4
    for s in sales:
        for it in s.get("items", []):
            ws2.cell(row=rr, column=1, value=s["id"]).alignment = _CENTER
            ws2.cell(row=rr, column=2, value=s["created_at"][:16])
            ws2.cell(row=rr, column=3, value=it.get("name", ""))
            ws2.cell(row=rr, column=4, value=it.get("qty", 0))
            ws2.cell(row=rr, column=5, value=it.get("unit", "dona"))
            cn = ws2.cell(row=rr, column=6, value=float(it.get("price", 0)))
            cn.number_format = '#,##0'
            cu = ws2.cell(row=rr, column=7, value=float(it.get("price_usd", 0)))
            cu.number_format = '$#,##0.00'
            ct = ws2.cell(row=rr, column=8, value=float(it.get("total", 0)))
            ct.number_format = '#,##0'
            ctu = ws2.cell(row=rr, column=9, value=float(it.get("total_usd", 0)))
            ctu.number_format = '$#,##0.00'
            for col in range(1, 10):
                ws2.cell(row=rr, column=col).border = _BORDER
            rr += 1
    _autosize(ws2)

    suffix = month or "hammasi"
    fname = f"sotuvlar_{suffix}_{now_local().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(_OUTPUT_DIR, fname)
    wb.save(path)
    return path


# ─── Mijozlar ────────────────────────────────────────────────────────────────

async def export_clients() -> str:
    clients = await db.get_all_clients()
    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar"
    ws["A1"] = f"Mijozlar — {now_local().strftime('%Y-%m-%d')}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")

    _write_header(
        ws,
        ["#", "Mijoz ismi", "Telefon", "Telegram ID",
         "Qarzi USD", "Qarzi so'm", "Ro'yxatga olingan"],
        row=3
    )
    for i, c in enumerate(clients, 1):
        r = i + 3
        ws.cell(row=r, column=1, value=c["id"]).alignment = _CENTER
        ws.cell(row=r, column=2, value=c["shop_name"])
        ws.cell(row=r, column=3, value=c.get("phone", ""))
        ws.cell(row=r, column=4, value=c.get("telegram_id", "") or "")
        cdu = ws.cell(row=r, column=5, value=float(c.get("debt_usd", 0) or 0))
        cdu.number_format = '$#,##0.00'
        cdu.alignment = _RIGHT
        cd = ws.cell(row=r, column=6, value=float(c.get("debt", 0)))
        cd.number_format = '#,##0'
        cd.alignment = _RIGHT
        if (c.get("debt", 0) or 0) > 0:
            cd.fill = PatternFill("solid", fgColor="FCE4D6")
            cdu.fill = PatternFill("solid", fgColor="FCE4D6")
        ws.cell(row=r, column=7, value=c.get("created_at", "")[:10])
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = _BORDER
    _autosize(ws)

    fname = f"mijozlar_{now_local().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(_OUTPUT_DIR, fname)
    wb.save(path)
    return path


# ─── Qarzdorlar ──────────────────────────────────────────────────────────────

async def export_debtors() -> str:
    debtors = await db.get_debtors()
    wb = Workbook()
    ws = wb.active
    ws.title = "Qarzdorlar"
    ws["A1"] = f"Qarzdorlar — {now_local().strftime('%Y-%m-%d')}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")

    _write_header(
        ws,
        ["#", "Mijoz", "Telefon", "Qarzi USD", "Qarzi so'm", "Ro'yxatga olingan"],
        row=3
    )
    total_sum = 0.0
    total_usd = 0.0
    for i, c in enumerate(debtors, 1):
        r = i + 3
        ws.cell(row=r, column=1, value=i).alignment = _CENTER
        ws.cell(row=r, column=2, value=c["shop_name"])
        ws.cell(row=r, column=3, value=c.get("phone", ""))
        cdu = ws.cell(row=r, column=4, value=float(c.get("debt_usd", 0) or 0))
        cdu.number_format = '$#,##0.00'
        cdu.alignment = _RIGHT
        cd = ws.cell(row=r, column=5, value=float(c.get("debt", 0)))
        cd.number_format = '#,##0'
        cd.alignment = _RIGHT
        ws.cell(row=r, column=6, value=c.get("created_at", "")[:10])
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = _BORDER
        total_sum += c.get("debt", 0) or 0
        total_usd += c.get("debt_usd", 0) or 0

    last = len(debtors) + 4
    ws.cell(row=last, column=3, value="JAMI:").font = Font(bold=True)
    cu = ws.cell(row=last, column=4, value=total_usd)
    cu.font = Font(bold=True); cu.number_format = '$#,##0.00'
    cu.fill = PatternFill("solid", fgColor="FCE4D6")
    cs = ws.cell(row=last, column=5, value=total_sum)
    cs.font = Font(bold=True); cs.number_format = '#,##0'
    cs.fill = PatternFill("solid", fgColor="FCE4D6")
    _autosize(ws)

    fname = f"qarzdorlar_{now_local().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(_OUTPUT_DIR, fname)
    wb.save(path)
    return path


# ─── Mahsulotlar (ombor holati) ──────────────────────────────────────────────

async def export_products() -> str:
    prods = await db.get_all_products(active_only=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ombor"
    ws["A1"] = f"Ombor holati — {now_local().strftime('%Y-%m-%d')}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:G1")

    _write_header(
        ws,
        ["ID", "Nomi", "Tannarx (USD)", "Tannarx (so'm)",
         "Sotish (USD)", "Sotish (so'm)", "Optom (USD)", "Optom (so'm)",
         "Qoldiq", "Birlik",
         "Ombor qiymati (USD)", "Ombor qiymati (so'm)"],
        row=3
    )
    total_value_sum = 0.0
    total_value_usd = 0.0
    for i, p in enumerate(prods, 1):
        r = i + 3
        qty = float(p.get("qty", 0) or 0)
        value_sum = qty * float(p.get("cost_price", 0) or 0)
        value_usd = qty * float(p.get("cost_price_usd", 0) or 0)
        ws.cell(row=r, column=1, value=p["id"]).alignment = _CENTER
        ws.cell(row=r, column=2, value=p["name"])
        c3 = ws.cell(row=r, column=3, value=float(p.get("cost_price_usd", 0) or 0))
        c3.number_format = '$#,##0.00'
        c4 = ws.cell(row=r, column=4, value=float(p.get("cost_price", 0)))
        c4.number_format = '#,##0'
        c5 = ws.cell(row=r, column=5, value=float(p.get("sell_price_usd", 0) or 0))
        c5.number_format = '$#,##0.00'
        c6 = ws.cell(row=r, column=6, value=float(p.get("sell_price", 0)))
        c6.number_format = '#,##0'
        c7 = ws.cell(row=r, column=7, value=float(p.get("wholesale_price_usd", 0) or 0))
        c7.number_format = '$#,##0.00'
        c8 = ws.cell(row=r, column=8, value=float(p.get("wholesale_price", 0)))
        c8.number_format = '#,##0'
        c9 = ws.cell(row=r, column=9, value=qty)
        c9.alignment = _RIGHT
        if qty <= 5:
            c9.fill = PatternFill("solid", fgColor="FCE4D6")
        ws.cell(row=r, column=10, value=p.get("unit", "dona")).alignment = _CENTER
        cvu = ws.cell(row=r, column=11, value=float(value_usd))
        cvu.number_format = '$#,##0.00'
        cvu.alignment = _RIGHT
        cv = ws.cell(row=r, column=12, value=float(value_sum))
        cv.number_format = '#,##0'
        cv.alignment = _RIGHT
        for col in range(1, 13):
            ws.cell(row=r, column=col).border = _BORDER
        total_value_sum += value_sum
        total_value_usd += value_usd

    last = len(prods) + 4
    ws.cell(row=last, column=10, value="JAMI:").font = Font(bold=True)
    cu = ws.cell(row=last, column=11, value=total_value_usd)
    cu.font = Font(bold=True); cu.number_format = '$#,##0.00'
    cu.fill = PatternFill("solid", fgColor="D9E1F2")
    cs = ws.cell(row=last, column=12, value=total_value_sum)
    cs.font = Font(bold=True); cs.number_format = '#,##0'
    cs.fill = PatternFill("solid", fgColor="D9E1F2")
    _autosize(ws)

    fname = f"ombor_{now_local().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(_OUTPUT_DIR, fname)
    wb.save(path)
    return path
